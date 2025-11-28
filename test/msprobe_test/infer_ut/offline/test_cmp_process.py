# Copyright (c) 2025-2025 Huawei Technologies Co., Ltd.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os, sys
import shutil
import unittest
from unittest.mock import Mock
from unittest.mock import patch, MagicMock, mock_open

import pandas as pd
import openpyxl
import pytest


from msprobe.infer.offline.compare.msquickcmp.cmp_process import _generate_golden_data_model, \
    _correct_the_wrong_order, _check_output_node_name_mapping, _get_single_csv_in_folder, _read_and_process_csv, \
    _write_csv, _append_column_to_csv, cmp_process, run_om_model_compare, csv_sum, _get_model_output_node_name_list, \
    _find_previous_node
from msprobe.infer.offline.compare.msquickcmp.common.utils import AccuracyCompareException
from msprobe.infer.offline.compare.msquickcmp.npu.npu_dump_data import NpuDumpData


class TestGenerateGoldenDataModel(unittest.TestCase):
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.onnx'))
    @patch('msprobe.infer.offline.compare.msquickcmp.onnx_model.onnx_dump_data.OnnxDumpData')
    def test_generate_onnx_model(self, mock_onnx_class, _):
        args = MagicMock()
        args.model_path = 'model.onnx'
        instance = MagicMock()
        mock_onnx_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'npu_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.onnx')

    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.om'))
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.NpuDumpData')
    def test_generate_om_model(self, mock_npu_class, _):
        args = MagicMock()
        args.model_path = 'model.om'
        instance = MagicMock()
        mock_npu_class.return_value = instance

        result, ext = _generate_golden_data_model(args, 'npu_path')
        self.assertEqual(result, instance)
        self.assertEqual(ext, '.om')

    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.get_model_name_and_extension', return_value=('model', '.unsupported'))
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger')
    def test_generate_unsupported_model(self, mock_logger, _):
        args = MagicMock()
        args.model_path = 'model.unsupported'

        with self.assertRaises(AccuracyCompareException):
            _generate_golden_data_model(args, 'npu_path')


class TestCorrectTheWrongOrder(unittest.TestCase):
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger')
    def test_correct_order(self, mock_logger):
        data = {0: 'a', 1: 'b'}
        _correct_the_wrong_order(0, 1, data)
        self.assertEqual(data[0], 'b')
        self.assertEqual(data[1], 'a')


class TestCheckOutputNodeNameMapping(unittest.TestCase):
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger')
    def test_check_output_node_name_mapping_match(self, mock_logger):
        output_node = {0: 'abc/xyz:0'}
        golden_info = {0: 'abc_xyz.0.out'}
        _check_output_node_name_mapping(output_node, golden_info)

    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger.warning')
    def test_check_output_node_name_mapping_no_match(self, mock_warn):
        output_node = {0: 'notmatch'}
        golden_info = {1: 'no_match_file.out'}
        _check_output_node_name_mapping(output_node, golden_info)
        mock_warn.assert_called()


class TestGetSingleCsvInFolder(unittest.TestCase):
    @patch('os.listdir', return_value=['data.csv'])
    def test_get_csv_found(self, _):
        path = '/some/path'
        result = _get_single_csv_in_folder(path)
        self.assertEqual(result, '/some/path/data.csv')

    @patch('os.listdir', return_value=['no_csv.txt'])
    def test_get_csv_not_found(self, _):
        with self.assertRaises(IOError):
            _get_single_csv_in_folder('/some/dir')


class TestReadAndProcessCSV(unittest.TestCase):
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.ms_open', new_callable=mock_open, read_data='a,b\n1,2')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.Rule.input_file')
    def test_read_and_process_csv(self, mock_rule, mock_file):
        mock_rule.return_value.check.return_value = True

        process_func = MagicMock(return_value=[['header1', 'header2'], ['val1', 'val2']])
        result = _read_and_process_csv('fake.csv', process_func, ['node'])
        self.assertTrue(isinstance(result, list))
        self.assertEqual(result[0], ['header1', 'header2'])


class TestWriteCsv(unittest.TestCase):
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.ms_open', new_callable=mock_open)
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.csv.writer')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.sanitize_csv_value', side_effect=lambda x: x)
    def test_write_csv(self, mock_sanitize, mock_csv_writer, mock_file):
        writer = MagicMock()
        mock_csv_writer.return_value = writer

        _write_csv('path.csv', [['a', 'b'], ['1', '2']])
        self.assertEqual(writer.writerows.call_count, 1)


@pytest.fixture(scope="function")
def import_cmp_process():
    backup = {}
    for mod in ['acl', 'msquickcmp.cmp_process']:
        if mod in sys.modules:
            backup[mod] = sys.modules[mod]
    mock_acl = MagicMock()
    sys.modules['acl'] = mock_acl
    functions = {
        "csv_sum": csv_sum,
        "_get_model_output_node_name_list": _get_model_output_node_name_list,
        "_find_previous_node": _find_previous_node
    }
    yield functions

    for mod, module_obj in backup.items():
        sys.modules[mod] = module_obj
    for mod in ['acl', 'msquickcmp.cmp_process']:
        if mod not in backup and mod in sys.modules:
            del sys.modules[mod]


@pytest.fixture(scope="function")
def generate_fake_path():
    cur_dir = os.path.dirname(os.path.realpath(__file__))
    resource_dir = os.path.join(cur_dir, 'test_csv_sum', "2023072009")

    os.makedirs(resource_dir, 0o750, exist_ok=True)
    os.makedirs(os.path.join(resource_dir, "images-2_3_638_640"), 0o750, exist_ok=True)
    os.makedirs(os.path.join(resource_dir, "images-2_3_640_640"), 0o750, exist_ok=True)

    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df1.to_csv(os.path.join(resource_dir, "images-2_3_638_640", "file1.csv"), index=False)

    df2 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df2.to_csv(os.path.join(resource_dir, "images-2_3_640_640", "file2.csv"), index=False)

    with pd.ExcelWriter(os.path.join(cur_dir, "test_csv_sum", "expected_output.xlsx")) as writer:
        df1.to_excel(writer, sheet_name='images-2_3_638_640', index=False)
        df2.to_excel(writer, sheet_name='images-2_3_640_640', index=False)
    yield resource_dir

    shutil.rmtree(os.path.join(cur_dir, "test_csv_sum"))


def test_csv_sum_given_path_when_valid_then_pass(import_cmp_process, generate_fake_path):
    csv_sum = import_cmp_process["csv_sum"]
    csv_sum(generate_fake_path)
    expected_output = openpyxl.load_workbook(os.path.join(generate_fake_path, '..', 'expected_output.xlsx'))
    result_summary = openpyxl.load_workbook(os.path.join(generate_fake_path, 'result_summary.xlsx'))

    sheets1 = result_summary.sheetnames
    sheets2 = expected_output.sheetnames

    assert len(sheets1) == len(sheets2)

    for sheet_name in sheets1:
        sheet1 = result_summary[sheet_name]
        sheet2 = expected_output[sheet_name]

        assert sheet1.max_row == sheet2.max_row
        assert sheet1.max_column == sheet2.max_column

        for row in range(1, sheet1.max_row + 1):
            for col in range(1, sheet1.max_column + 1):
                assert sheet1.cell(row=row, column=col).value == sheet2.cell(row=row, column=col).value
        

class MockSession:
    def get_outputs(self):
        return [MockNode(name="output_1", outputs=["net_output"]), MockNode(name="output_2", outputs=["net_output"])]


class MockGraph:
    def __init__(self, nodes):
        self.node = nodes


class MockNode:
    def __init__(self, name, outputs):
        self.name = name
        self.output = outputs


def test_find_previous_node(import_cmp_process):
    _find_previous_node = import_cmp_process["_find_previous_node"]
    # 创建模拟的节点
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    graph = MockGraph(nodes=[node1, node2])

    # 测试找到前一个节点
    assert _find_previous_node(graph, "output_1") == "node1"
    assert _find_previous_node(graph, "output_2") == "node2"

    # 测试未找到前一个节点
    assert _find_previous_node(graph, "output_3") is None


def test_get_model_output_node_name_list(import_cmp_process):
    _get_model_output_node_name_list = import_cmp_process["_get_model_output_node_name_list"]
    # 创建模拟的会话和模型
    session = MockSession()
    node1 = MockNode(name="node1", outputs=["output_1"])
    node2 = MockNode(name="node2", outputs=["output_2"])
    origin_model = Mock(graph=MockGraph(nodes=[node1, node2]))
    # 测试正常情况
    result = _get_model_output_node_name_list(session, origin_model)
    assert result == ["node1", "node2"], f"Expected ['node1', 'node2'], but got {result}"

    # 测试找不到前一个节点的情况
    node3 = MockNode(name="node3", outputs=["output_3"])
    origin_model.graph.node = [node3]  # 修改模型节点
    result = _get_model_output_node_name_list(session, origin_model)
    assert result is None, f"Expected None, but got {result}"


class TestDumpCmpFunctions(unittest.TestCase):
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process.check_and_run")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process.os.path.realpath", return_value="/real/path")
    def test_cmp_process_dump(self, mock_realpath, mock_check_and_run):
        args = MagicMock()
        args.target_path = None
        args.golden_path = None
        args.ops_json = None
        cmp_process(args)
        mock_check_and_run.assert_called_once()


class TestAppendColumnToCSV(unittest.TestCase):

    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._write_csv")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._read_and_process_csv")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._get_single_csv_in_folder")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._process_is_npu_and_is_precision_error_ops")
    def test_append_column_to_csv_with_node_output_show_list(self, mock_process_func, mock_get_csv, mock_read_process, mock_write):
        mock_get_csv.return_value = "/fake/path/file.csv"
        mock_read_process.return_value = [["row1"], ["row2"]]
        node_output_show_list = ["node1", "node2"]
        _append_column_to_csv("/some/path", node_output_show_list)
        mock_get_csv.assert_called_once_with("/some/path")
        mock_read_process.assert_called_once_with("/fake/path/file.csv", mock_process_func, node_output_show_list)
        mock_write.assert_called_once_with("/fake/path/file.csv", [["row1"], ["row2"]])

    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._write_csv")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._read_and_process_csv")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._get_single_csv_in_folder")
    @patch("msprobe.infer.offline.compare.msquickcmp.cmp_process._process_is_npu_and_is_precision_error_ops")
    def test_append_column_to_csv_with_none_node_output_show_list(self, mock_process_func, mock_get_csv, mock_read_process, mock_write):
        mock_get_csv.return_value = "/fake/path/file.csv"
        mock_read_process.return_value = [["row1"]]
        _append_column_to_csv("/some/path", None)
        mock_get_csv.assert_called_once_with("/some/path")
        mock_read_process.assert_called_once_with("/fake/path/file.csv", mock_process_func, [])
        mock_write.assert_called_once_with("/fake/path/file.csv", [["row1"]])


class TestRunOmModelCompare(unittest.TestCase):

    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process._append_column_to_csv')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process._get_model_output_node_name_list')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process._check_output_node_name_mapping')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.NetCompare')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.handle_ground_truth_files')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process._generate_golden_data_model')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.NpuDumpData')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.OmParser')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.atc_utils.convert_model_to_json')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger')
    def test_run_om_model_compare_basic(self, mock_logger,
                                        mock_convert_model_to_json,
                                        mock_OmParser,
                                        mock_NpuDumpData,
                                        mock_generate_golden_data_model,
                                        mock_handle_ground_truth_files,
                                        mock_NetCompare,
                                        mock_check_output_node_name_mapping,
                                        mock_get_model_output_node_name_list,
                                        mock_append_column_to_csv):

        args = MagicMock()
        args.cann_path = "/fake/cann"
        args.target_path = "/fake/offline.om"
        args.output_path = "/fake/out"
        args.golden_path = "/fake/model.om"
        args.dump = False

        mock_convert_model_to_json.side_effect = lambda cann, model, out: "/fake/json_path"

        omparser_instance = MagicMock()
        omparser_instance.get_aipp_config_content.return_value = True
        mock_OmParser.return_value = omparser_instance

        npu_dump_instance = MagicMock()
        mock_NpuDumpData.return_value = npu_dump_instance
        mock_NpuDumpData.__bases__ = (NpuDumpData,)
        mock_NpuDumpData.__class__ = type(NpuDumpData)

        npu_dump_instance.generate_inputs_data.return_value = None

        npu_dump_instance.generate_dump_data.return_value = ("/fake/npu_dump_path", "/fake/npu_net_output_path")

        npu_dump_instance.get_expect_output_name.return_value = ["output_node"]

        golden_dump_mock = MagicMock()
        mock_generate_golden_data_model.return_value = (golden_dump_mock, ".onnx")

        golden_dump_mock.generate_inputs_data.return_value = None
        golden_dump_mock.generate_dump_data.return_value = "/fake/golden_dump_path"
        golden_dump_mock.get_net_output_info.return_value = {"some": "info"}

        mock_handle_ground_truth_files.return_value = None

        net_compare_instance = MagicMock()
        mock_NetCompare.return_value = net_compare_instance
        net_compare_instance.net_output_compare.return_value = None

        mock_check_output_node_name_mapping.return_value = None

        mock_get_model_output_node_name_list.return_value = ["node1", "node2"]

        mock_append_column_to_csv.return_value = None

        run_om_model_compare(args)

        mock_convert_model_to_json.assert_called()
        mock_OmParser.assert_called()
        npu_dump_instance.generate_inputs_data.assert_called()
        npu_dump_instance.generate_dump_data.assert_called()
        mock_generate_golden_data_model.assert_called()
        mock_handle_ground_truth_files.assert_called()
        mock_NetCompare.assert_called()
        mock_check_output_node_name_mapping.assert_called()
        mock_get_model_output_node_name_list.assert_called()
        mock_append_column_to_csv.assert_called()

    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.OmParser')
    @patch('msprobe.infer.offline.compare.msquickcmp.cmp_process.utils.logger')
    def test_raise_when_fusion_switch_file_and_aipp(self, mock_logger, mock_OmParser):
        args = MagicMock()
        args.cann_path = "a"
        args.target_path = "b"
        args.output_path = "c"
        args.golden_path = "model.om"
        args.dump = False

        omparser_instance = MagicMock()
        omparser_instance.get_aipp_config_content.return_value = True
        mock_OmParser.return_value = omparser_instance
        with self.assertRaises(AccuracyCompareException):
            run_om_model_compare(args)
        mock_logger.error.assert_called()
